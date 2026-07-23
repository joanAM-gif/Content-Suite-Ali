"""
Modulo III: Governance & Multimodal Audit.

Dos responsabilidades separadas, una por cada aprobador:

- Aprobador A revisa el TEXTO generado en el Modulo II y lo aprueba o
  rechaza (GET /review/pending, PATCH /review/{id}).
- Aprobador B sube una IMAGEN y el sistema la audita contra el manual de
  marca usando el modelo de vision de Google AI Studio (POST /audit-image).

`audit_product_image` esta decorado con @observe(): es el trace raiz que
agrupa el retrieval del RAG (Modulo I) y la generacion de vision
(Gemini), y en Langfuse se ve directamente cuanto tomo cada parte y el
total - la "auditoria de procesos" que pide el Modulo IV.
"""
import io
from datetime import datetime

from fastapi import APIRouter, File, Form, HTTPException, UploadFile  # type: ignore[reportMissingImports]
from fastapi.responses import StreamingResponse  # type: ignore[reportMissingImports]

from app.database import get_connection
from app.schemas import ContentPieceOut, ImageAuditResponse, MetricsResponse, ReviewUpdate
from app.services import notifications          # <- nueva
from app.services.observability import observe  # type: ignore[reportAttributeAccessIssue]
from app.services.rag import retrieve_brand_context
from app.services.vision import audit_image


router = APIRouter(tags=["Modulo III - Governance & Multimodal Audit"])


def _iso(value) -> str | None:
    """Convierte un timestamp de psycopg2 a string ISO para el schema (o None)."""
    return value.isoformat() if value is not None else None


@router.get("/review/pending", response_model=list[ContentPieceOut])
def list_pending_reviews() -> list[ContentPieceOut]:
    """Piezas de contenido en estado 'pendiente', para la bandeja del Aprobador A."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, product, content_type, content, status, reviewer_note,
                           created_at, updated_at
                    FROM content_pieces
                    WHERE status = 'pendiente'
                    ORDER BY created_at DESC
                    """
                )
                rows = cur.fetchall()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error consultando piezas pendientes: {exc}") from exc

    return [
        ContentPieceOut(
            id=row[0], product=row[1], content_type=row[2],
            content=row[3], status=row[4], reviewer_note=row[5],
            created_at=_iso(row[6]), updated_at=_iso(row[7]),
        )
        for row in rows
    ]


@router.get("/review/history", response_model=list[ContentPieceOut])
def list_review_history() -> list[ContentPieceOut]:
    """
    Bitacora completa: TODAS las piezas de contenido (pendiente, aprobado y
    rechazado), con quien las reviso (nota) y cuando, ordenadas por ultima
    actualizacion. Es la evidencia de trazabilidad para "Gobernanza de Datos".
    """
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, product, content_type, content, status, reviewer_note,
                           created_at, updated_at
                    FROM content_pieces
                    ORDER BY updated_at DESC
                    """
                )
                rows = cur.fetchall()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error consultando el historial: {exc}") from exc

    return [
        ContentPieceOut(
            id=row[0], product=row[1], content_type=row[2],
            content=row[3], status=row[4], reviewer_note=row[5],
            created_at=_iso(row[6]), updated_at=_iso(row[7]),
        )
        for row in rows
    ]


@router.get("/metrics", response_model=MetricsResponse)
def get_metrics() -> MetricsResponse:
    """Conteos agregados para el panel de metricas (extra sobre lo pedido en el reto)."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT status, count(*) FROM content_pieces GROUP BY status")
                status_counts = dict(cur.fetchall())

                cur.execute("SELECT cumple, count(*) FROM image_audits GROUP BY cumple")
                audit_counts = dict(cur.fetchall())
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error calculando metricas: {exc}") from exc

    pendiente = status_counts.get("pendiente", 0)
    aprobado = status_counts.get("aprobado", 0)
    rechazado = status_counts.get("rechazado", 0)
    cumple = audit_counts.get(True, 0)
    no_cumple = audit_counts.get(False, 0)

    return MetricsResponse(
        total_generado=pendiente + aprobado + rechazado,
        pendiente=pendiente,
        aprobado=aprobado,
        rechazado=rechazado,
        total_auditorias=cumple + no_cumple,
        auditorias_cumple=cumple,
        auditorias_no_cumple=no_cumple,
    )

@router.get("/metrics/export")
def export_metrics() -> StreamingResponse:
    """
    Exporta el detalle completo de metricas a un archivo Excel (.xlsx):
    extra sobre lo pedido en el reto, para que el panel de metricas no se
    quede solo en un resumen en pantalla sino que se pueda llevar a un
    reporte fuera de la aplicacion.

    El archivo tiene 3 hojas: un resumen agregado (igual a GET /metrics),
    el detalle de todas las piezas de contenido, y el detalle de todas
    las auditorias de imagen.
    """
    # openpyxl is an optional runtime dependency used only by this export.
    # The backend environment may not expose its source package to Pylance.
    from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
    from openpyxl.styles import Font, PatternFill  # type: ignore[reportMissingModuleSource]

    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT status, count(*) FROM content_pieces GROUP BY status")
                status_counts = dict(cur.fetchall())

                cur.execute("SELECT cumple, count(*) FROM image_audits GROUP BY cumple")
                audit_counts = dict(cur.fetchall())

                cur.execute(
                    """
                    SELECT id, product, content_type, status, reviewer_note, created_at, updated_at
                    FROM content_pieces
                    ORDER BY created_at DESC
                    """
                )
                pieces = cur.fetchall()

                cur.execute(
                    "SELECT id, product, cumple, razon, created_at FROM image_audits ORDER BY created_at DESC"
                )
                audits = cur.fetchall()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error generando el export de metricas: {exc}") from exc

    pendiente = status_counts.get("pendiente", 0)
    aprobado = status_counts.get("aprobado", 0)
    rechazado = status_counts.get("rechazado", 0)
    cumple = audit_counts.get(True, 0)
    no_cumple = audit_counts.get(False, 0)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B1A26", end_color="1B1A26", fill_type="solid")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill

    def autosize(ws):
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 60)

    wb = Workbook()

    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Resumen"
    ws1.append(["Metrica", "Valor"])
    ws1.append(["Total contenido generado", pendiente + aprobado + rechazado])
    ws1.append(["Pendiente", pendiente])
    ws1.append(["Aprobado", aprobado])
    ws1.append(["Rechazado", rechazado])
    ws1.append(["Total auditorias de imagen", cumple + no_cumple])
    ws1.append(["Auditorias que cumplen", cumple])
    ws1.append(["Auditorias que no cumplen", no_cumple])
    ws1.append(["Exportado el", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    style_header(ws1)
    autosize(ws1)

    ws2 = wb.create_sheet("Piezas de contenido")
    ws2.append(["ID", "Producto", "Tipo", "Estado", "Nota del revisor", "Creado", "Actualizado"])
    for row in pieces:
        ws2.append([row[0], row[1], row[2], row[3], row[4] or "", _iso(row[5]) or "", _iso(row[6]) or ""])
    style_header(ws2)
    autosize(ws2)

    ws3 = wb.create_sheet("Auditorias de imagen")
    ws3.append(["ID", "Producto", "Cumple", "Razon", "Fecha"])
    for row in audits:
        ws3.append([row[0], row[1], "Si" if row[2] else "No", row[3], _iso(row[4]) or ""])
    style_header(ws3)
    autosize(ws3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"content-suite-metricas-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/review/{piece_id}", response_model=ContentPieceOut)
def review_content_piece(piece_id: int, payload: ReviewUpdate) -> ContentPieceOut:
    """Aprobador A aprueba o rechaza una pieza de contenido generada en el Modulo II."""
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE content_pieces
                    SET status = %s, reviewer_note = %s, updated_at = now()
                    WHERE id = %s
                    RETURNING id, product, content_type, content, status, reviewer_note,
                              created_at, updated_at
                    """,
                    (payload.status, payload.reviewer_note, piece_id),
                )
                row = cur.fetchone()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error actualizando la pieza de contenido: {exc}") from exc

    if row is None:
        raise HTTPException(status_code=404, detail=f"No existe la pieza de contenido {piece_id}.")

    piece = ContentPieceOut(
        id=row[0], product=row[1], content_type=row[2],
        content=row[3], status=row[4], reviewer_note=row[5],
        created_at=_iso(row[6]), updated_at=_iso(row[7]),
    )

    # Notificacion por email (extra sobre lo pedido en el reto). Best-effort:
    # si falla, no debe romper la respuesta de la API (ver services/notifications.py).
    try:
        notifications.notify_review_status(piece.id, piece.product, piece.status, piece.reviewer_note)
    except Exception:  # noqa: BLE001
        pass

    return piece


@router.post("/audit-image", response_model=ImageAuditResponse)
@observe(name="POST /audit-image")
async def audit_product_image(
    producto: str = Form(...),
    image: UploadFile = File(...),
) -> ImageAuditResponse:
    """Aprobador B sube una imagen que se audita contra el manual de marca del producto."""
    try:
        brand_chunks = retrieve_brand_context(
            product=producto,
            query=f"reglas visuales y de tono de marca de {producto}",
            top_k=8,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error recuperando el manual de marca: {exc}") from exc

    if not brand_chunks:
        raise HTTPException(
            status_code=404,
            detail=f"No hay manual de marca indexado para '{producto}'. Genera uno primero con POST /brand.",
        )

    image_bytes = await image.read()

    try:
        verdict = audit_image(
            product=producto,
            brand_rules=[chunk["content"] for chunk in brand_chunks],
            image_bytes=image_bytes,
            mime_type=image.content_type or "image/jpeg",
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Error auditando la imagen con Gemini: {exc}") from exc

    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO image_audits (product, cumple, razon)
                    VALUES (%s, %s, %s)
                    RETURNING id
                    """,
                    (producto, verdict["cumple"], verdict["razon"]),
                )
                row = cur.fetchone()
                if row is None:
                    raise HTTPException(status_code=500, detail="Error guardando la auditoria: no se devolvió el id")
                audit_id = row[0]
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Error guardando la auditoria: {exc}") from exc

    # Notificacion por email (extra sobre lo pedido en el reto). Best-effort.
    try:
        notifications.notify_image_audit(audit_id, producto, verdict["cumple"], verdict["razon"])
    except Exception:  # noqa: BLE001
        pass

    return ImageAuditResponse(
        id=audit_id, producto=producto,
        cumple=verdict["cumple"], razon=verdict["razon"],
    )